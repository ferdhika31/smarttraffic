# -*- coding: utf-8 -*-
# @Author			: Ferdhika
# @Date				: 2017-04-16 14:56:41

from bottle import response

import json

# Library twython twitter
from twython import Twython

# Library excel
from openpyxl import load_workbook
from pymongo import MongoClient

# Dari pebahasa (https://github.com/pebbie/pebahasa)
from hmmtagger import MainTagger
import suku
from tokenization import *
from termextract import *

'''
GLOBAL VARIABLE
'''
# buat koneksi untuk menghubungkan komunikasi dengan Mongo DB
client = MongoClient('localhost:27017')
db = client.ProyekData # Nama database

mt = None

# Function ambil tweet
def getTweet(username, jumlah):
	# token dari apps developer twitter
	access_token_key = '288561401-XRj4ltLCk8eOOysLBUFJBIrhI9gYLdlcSiOwRfaQ'
	access_token_secret = 'nhYr25lhZlmsI1ZvRlR954avggp0nhMwY3zMpFFKC4w8K'
	consumer_key = 'v6u60V22UcI7gswkOEhXetVrT'
	consumer_secret = 'hM6h2TlBLURw4RqWXVFWfYZv3YSETMxUAKvK1U0ZLvlSOiiipl'

	twitter = Twython(consumer_key, consumer_secret,access_token_key, access_token_secret)
	# dishub_bdgkab
	return twitter.get_user_timeline(screen_name=username,count=jumlah)

# inisiasi tagger
def init_tag():
    global mt
    if mt is None:
        mt = MainTagger("resource/Lexicon.trn", "resource/Ngram.trn", 0, 3, 3, 0, 0, False, 0.2, 0, 500.0, 1)

# inisiasi kamus
def init_kamus():
	'''
	Kamus Normalisasi
	'''
	#read workbook to get data
	sheet = 'Sheet1'
	wb = load_workbook(filename = 'resource/list_normalisasi.xlsx')
	ws = wb.get_sheet_by_name(name = sheet)
	sheet_ranges = wb[sheet]

	res = []

	#Iterate through worksheet and insert cell contents to mongo
	n =2
	for row in ws.iter_rows('A{}:A{}'.format(ws.min_row,ws.max_row-1), row_offset=1):
		for cell in row:
			data = db.KamusNormalisasi.find_one({ 'kata': cell.value })
			# Cek kata, kalo belum ada di mongo db baru di tambahin
			if not data:
				db.KamusNormalisasi.insert_one({
					'kata'		: cell.value,
					'maksud'	: sheet_ranges['B'+str(n)].value,
					'sentiment'	: sheet_ranges['C'+str(n)].value,
				})
		n=n+1

# inisiasi data jalan
def init_data_jalan():
	'''
	Kamus Normalisasi
	'''
	#read workbook to get data
	sheet = 'Sheet4'
	wb = load_workbook(filename = 'resource/list_data_jalan.xlsx')
	ws = wb.get_sheet_by_name(name = sheet)
	sheet_ranges = wb[sheet]

	res = []

	#Iterate through worksheet and insert cell contents to mongo
	n =2
	for row in ws.iter_rows('A{}:A{}'.format(ws.min_row,ws.max_row-1), row_offset=1):
		for cell in row:
			data = db.DataJalan.find_one({ 'kata': cell.value })
			# Cek kata, kalo belum ada di mongo db baru di tambahin
			if not data:
				db.DataJalan.insert_one({
					'kata'	: cell.value,
					'jadi'	: sheet_ranges['B'+str(n)].value
				})
		n=n+1

def jalan(teks):

	data = db.DataJalan.find()

	for x in data:
		teks = teks.replace(x['kata'], x['jadi'], 4)

	return teks

# Function tagger
def tag(teks):
	response.content_type = 'application/json'
	lines = teks.strip().split("\n")
	result = []
	try:
		init_tag()
		for l in lines:
			if len(l) == 0: continue
			out = sentence_extraction(cleaning(l))
			for o in out:
				strtag = " ".join(tokenisasi_kalimat(o)).strip()
				result += [" ".join(mt.taggingStr(strtag))]
	except:
		return "Error Exception"
	return "\n".join(result)

# Function sentence
def sentence(tweet):
	response.content_type = 'application/json'
 
	res = remove_mention(tweet)
 
	res = res.split(' ')
	tag = []    
	num = 0
	tmp=1
	
	kirim = []
	tampil = []

	cdp = False
	nnp = False
	sampai=False

	for kata in res:
		tag.append(kata.split('/'))

		if tag[num][1] == 'CDP':
			cdp=True
			# print('Waktu : ' + tag[num][0])
 			kirim.append(tag[num][0])
		if tag[num][1] == 'NNP':
			nnp=True
			if tmp==1:
				# print('Dari : ' + tag[num][0])
				kirim.append(tag[num][0])
			else:
				sampai=True
				# print('Sampai : ' + tag[num][0])
				kirim.append(tag[num][0])
			tmp=tmp+1
 
		num = num + 1

	if cdp and nnp:
		if sampai:
			tampil.append({
				'waktu' : kirim[0],
				'dari' : kirim[1],
				'sampai' : kirim[2]   
			});
		else:
			tampil.append({
				'waktu' : kirim[0],
				'dari' : kirim[1] 
			});

	tampil = json.dumps(tampil, sort_keys=True, indent=5)

	return tampil

# Function remove link url
def remove_urls(vTEXT):
	vTEXT = re.sub(r'(https|http)?:\/\/(\w|\.|\/|\?|\=|\&|\%)*\b', '', vTEXT, flags=re.MULTILINE)
	return(vTEXT)

# Function remove mention
def remove_mention(vTEXT):
	vTEXT = re.sub(r'(@)(?<=@)\w+', '', vTEXT, flags=re.MULTILINE)
	return(vTEXT)